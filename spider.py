from selenium import webdriver
from selenium.common import TimeoutException, NoSuchElementException
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as ec

from bs4 import BeautifulSoup
from openpyxl import Workbook

options = webdriver.ChromeOptions()

options.add_argument("--disable-blink-features=AutomationControlled")
options.add_argument('window-size=1300,750')
options.add_argument('window-position=300,90')

# Exclude the collection of enable-automation switches
options.add_experimental_option("excludeSwitches", ["enable-automation"])

# Turn-off userAutomationExtension
options.add_experimental_option("useAutomationExtension", False)

# Setting the driver path and requesting a page
driver = webdriver.Chrome(options=options)

# Changing the property of the navigator value for webdriver to undefined
driver.execute_script("Object.defineProperty(navigator, 'webdriver', {get: () => undefined})")

driver.get('https://2gis.ae')

driver.implicitly_wait(10)
categories = driver.find_elements(By.CLASS_NAME, '_mq2eit')
categories_btn = categories[-2]
categories_btn.find_element(By.CLASS_NAME, '_1g22egm').click()

categories = driver.find_elements(By.CSS_SELECTOR, '._r47nf ._1g22egm')

data = []

for index, category in enumerate(categories[:-2]):
    if index == 0:
        category_name = str(category.text)
        category.click()
        sub_categories = driver.find_elements(By.CSS_SELECTOR, '._r47nf + ._r47nf ._1g22egm')
        sub_cat_names = [sub_cat.find_element(By.CSS_SELECTOR, 'span._lt317').text for sub_cat in sub_categories[:-1]]
        sub_categories_links = [element.get_attribute('href') for element in sub_categories[:-1]]
        for index2, new_tab_link in enumerate(sub_categories_links):
            driver.execute_script(f"window.open('{new_tab_link}');")
            driver.switch_to.window(driver.window_handles[1])

            i = 0
            finish = False
            while not finish:

                WebDriverWait(driver, 10).until(ec.visibility_of_element_located((By.CSS_SELECTOR, 'div._zjunba a._1rehek')))
                companies = driver.find_elements(By.CSS_SELECTOR, 'div._zjunba a._1rehek')

                div = driver.find_element(By.CSS_SELECTOR, '._r47nf').get_attribute('outerHTML')
                soup = BeautifulSoup(div, 'html.parser')
                addresses_driver = soup.select('._klarpw ._1w9o2igt:nth-child(1)')
                business_names = soup.select('._1al0wlf span')
                addresses = [address.text.replace('\xa0', '').replace('\u200b', '') for address in addresses_driver]

                for index3, company in enumerate(companies):
                    driver.execute_script(f"window.open('{company.get_attribute('href')}');")
                    driver.switch_to.window(driver.window_handles[2])

                    try:
                        WebDriverWait(driver, 10).until(ec.presence_of_element_located((By.CSS_SELECTOR, "._9idr87 + div")))
                    except TimeoutException:
                        raise Exception("Failed to find business information")

                    business_source = driver.find_element(By.CSS_SELECTOR, '._9idr87 + div').get_attribute('outerHTML')

                    soup = BeautifulSoup(business_source, 'html.parser')

                    business_name = business_names[index3].text
                    try:
                        website = soup.select_one('span div ._1rehek').text
                        if website.strip().lower() == str('Consumer rights feedback').lower():
                            website = ''
                    except:
                        website = ''
                    try:
                        email = ''
                        strings = soup.select('._49kxlr ._2lcm958')
                        for string in strings:
                            if "@" in string.text:
                                email = string.text
                    except:
                        email = ''
                    address = addresses[index3]
                    media = soup.find_all('div', {'class': '_14uxmys'})
                    try:
                        phone_element = soup.select_one('div._b0ke8 a')
                        if phone_element:
                            phone = phone_element['href']
                        else:
                            phone = ''
                    except (AttributeError, TypeError) as e:
                        phone = ''
                    media_elements = {'Facebook': '', 'Instagram': '', 'Twitter': '', 'LinkedIn': '', 'YouTube': '', 'Tiktok': ''}
                    for element in media:
                        media_elements[element.text] = element.select_one('a._1rehek').get('href')

                    item = {'Category': category_name,
                            'sub-Category': sub_cat_names[index2],
                            'Business Name': business_name,
                            'website': website,
                            'Facebook': media_elements['Facebook'],
                            'Instagram': media_elements['Instagram'],
                            'Twitter': media_elements['Twitter'],
                            'LinkedIn': media_elements['LinkedIn'],
                            'YouTube': media_elements['YouTube'],
                            'Tiktok': media_elements['Tiktok'],
                            'Email': email,
                            'Contact No.': phone[4:],
                            'Address': address}

                    print(item)
                    data.append(item)
                    driver.close()
                    driver.switch_to.window(driver.window_handles[1])

                try:
                    next_page = driver.find_element(By.CSS_SELECTOR, 'div._n5hmn94 + div._n5hmn94')
                    driver.execute_script("arguments[0].scrollIntoView();", next_page)
                    driver.execute_script("arguments[0].click();", next_page)
                except NoSuchElementException as e:
                    try:
                        next_page = driver.find_element(By.CSS_SELECTOR, 'div._7q94tr + div._n5hmn94')
                        driver.execute_script("arguments[0].scrollIntoView();", next_page)
                        driver.execute_script("arguments[0].click();", next_page)
                    except NoSuchElementException as e:
                        finish = True
                i += 1
                print('entered: ' + str(i))
            driver.close()
            driver.switch_to.window(driver.window_handles[0])

print('\nExit Succeed')
driver.quit()

# Save into Excel
workbook = Workbook()
sheet = workbook.active
header_row = 1
for col, key in enumerate(data[0].keys(), start=1):
    sheet.cell(row=header_row, column=col, value=key)

data_row = 2
for item in data:
    for col, value in enumerate(item.values(), start=1):
        sheet.cell(row=data_row, column=col, value=value)
    data_row += 1

workbook.save('Eat Out.xlsx')
